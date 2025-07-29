import json
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

# Шлях до шаблону Excel (в форматі .xlsx)
TEMPLATE_PATH = "template_vfi.xlsx"

# Папка для збереження накладних
INVOICE_FOLDER = "invoices_vfi"

# Завантаження даних із файлів JSON
def load_data(customers_file, products_file, history_file, counter_file):
    with open(customers_file, "r", encoding="utf-8") as f:
        customers = json.load(f)
    with open(products_file, "r", encoding="utf-8") as f:
        products = json.load(f)
    with open(history_file, "r", encoding="utf-8") as f:
        history = json.load(f)
    with open(counter_file, "r", encoding="utf-8") as f:
        counter = json.load(f)
    return customers, products, history, counter

# Збереження історії та лічильника
def save_history_and_counter(history_file, history, counter_file, counter):
    with open(history_file, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=4)
    with open(counter_file, "w", encoding="utf-8") as f:
        json.dump(counter, f, ensure_ascii=False, indent=4)

# Генерація накладної Excel
def generate_invoice(client_name, invoice_number, items):
    if not os.path.exists(INVOICE_FOLDER):
        os.makedirs(INVOICE_FOLDER)
    filename = os.path.join(INVOICE_FOLDER, f"{client_name}_{invoice_number:04d}.xlsx")

    # Копіюємо шаблон
    shutil.copyfile(TEMPLATE_PATH, filename)

    wb = load_workbook(filename)
    ws = wb.active  # або конкретний лист, якщо треба: wb["Invoice"]

    # Заповнюємо дані в шаблоні
    ws["B2"] = f"{invoice_number:04d}"  # Номер накладної
    ws["B3"] = datetime.now().strftime("%d.%m.%Y")  # Дата
    ws["B4"] = client_name  # Ім'я клієнта

    start_row = 7  # Початок вставки товарів

    # Записуємо товари
    for idx, item in enumerate(items, start=start_row):
        ws[f"A{idx}"] = idx - start_row + 1  # Позиція
        ws[f"B{idx}"] = item["name"]  # Назва товару
        ws[f"C{idx}"] = item["quantity"]  # Кількість
        ws[f"D{idx}"] = item["unit_price"]  # Ціна за одиницю
        ws[f"E{idx}"] = item["quantity"] * item["unit_price"]  # Сума

    # Загальна сума
    total_row = start_row + len(items)
    ws[f"D{total_row}"] = "Всього:"
    ws[f"E{total_row}"] = f"=SUM(E{start_row}:E{total_row -1})"

    wb.save(filename)
    print(f"Накладна збережена у файл: {filename}")

# Основна функція роботи модуля
def run_module():
    # Файли для ВФІ
    customers_file = "customers_vfi.json"
    products_file = "products_vfi.json"
    history_file = "history_vfi.json"
    counter_file = "counter_vfi.json"

    customers, products, history, counter = load_data(customers_file, products_file, history_file, counter_file)

    while True:
        print("\nСписок клієнтів:")
        for idx, client in enumerate(customers, start=1):
            print(f"{idx}. {client}")

        client_choice = input("Введіть ім'я клієнта для створення накладної (або Enter для виходу): ").strip()
        if client_choice == "":
            break
        if client_choice not in customers:
            print("Клієнт не знайдений.")
            continue

        # Отримуємо номер накладної
        invoice_number = counter.get(client_choice, 0) + 1

        items = []
        while True:
            product_search = input("Назва товару (Enter для завершення): ").strip().lower()
            if product_search == "":
                break
            # Фільтруємо товари по введеному тексту
            filtered_products = [p for p in products if product_search in p["name"].lower()]
            if not filtered_products:
                print("Товар не знайдено.")
                continue
            for i, prod in enumerate(filtered_products, start=1):
                print(f"{i}. {prod['name']} (ціна: {prod['unit_price']})")
            selection = input("Виберіть номер товару: ").strip()
            if not selection.isdigit() or not (1 <= int(selection) <= len(filtered_products)):
                print("Некоректний вибір.")
                continue
            selected_product = filtered_products[int(selection) -1]

            quantity_str = input("Кількість: ").strip()
            if not quantity_str.isdigit() or int(quantity_str) <= 0:
                print("Некоректна кількість.")
                continue
            quantity = int(quantity_str)

            items.append({
                "name": selected_product["name"],
                "quantity": quantity,
                "unit_price": selected_product["unit_price"]
            })

        if not items:
            print("Накладна не створена (нема товарів).")
            continue

        # Генеруємо накладну
        generate_invoice(client_choice, invoice_number, items)

        # Оновлюємо історію та лічильник
        history.setdefault(client_choice, []).append({
            "invoice_number": invoice_number,
            "items": items,
            "date": datetime.now().strftime("%d.%m.%Y")
        })
        counter[client_choice] = invoice_number

        save_history_and_counter(history_file, history, counter_file, counter)
        print("Накладна успішно створена та збережена.\n")

if __name__ == "__main__":
    run_module()
