import os
import json
from datetime import datetime
from shutil import copyfile
from openpyxl import load_workbook
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt
from rich.panel import Panel
from rich.box import HEAVY
import matplotlib.pyplot as plt

console = Console()

DATA = {
    "ВФІ": {
        "customers": "customers_vfi.json",
        "products": "products_vfi.json",
        "history": "history_vfi.json",
        "counter": "counter_vfi.json",
        "template": "template_vfi.xlsx",
        "invoice_dir": "накладні_ВФІ"
    },
    "СМС": {
        "customers": "customers_sms.json",
        "products": "products_sms.json",
        "history": "history_sms.json",
        "counter": "counter_sms.json",
        "template": "template_sms.xlsx",
        "invoice_dir": "накладні_СМС"
    }
}

PRICE_TYPES = ["PK", "PM", "PSS", "PL", "PWS"]

def load_json(path):
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump({}, f)
    with open(path, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except:
            return {}

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def pause():
    input("\nНатисніть Enter для продовження...")

def print_clients(clients):
    table = Table(title="Клієнти", box=HEAVY)
    table.add_column("№")
    table.add_column("Клієнт")
    table.add_column("Тип ціни")
    for i, (name, price_type) in enumerate(clients.items(), 1):
        table.add_row(str(i), name, price_type)
    console.print(table)

def print_products(products):
    table = Table(title="Товари", box=HEAVY)
    table.add_column("№")
    table.add_column("Назва")
    for pt in PRICE_TYPES:
        table.add_column(pt)
    for i, (name, data) in enumerate(products.items(), 1):
        row = [str(i), name] + [str(data.get(pt, "")) for pt in PRICE_TYPES]
        table.add_row(*row)
    console.print(table)

def context_search(data, query):
    q = query.strip().lower()
    return [k for k in data if q in k.lower()]

def add_client(clients, path):
    name = Prompt.ask("Назва клієнта (Enter — скасувати)").strip()
    if not name:
        return
    price = Prompt.ask("Тип ціни", choices=PRICE_TYPES)
    clients[name] = price
    save_json(path, clients)
    console.print("[green]Клієнта додано[/green]")

def mass_add_products(products, path):
    price_type = Prompt.ask("Тип ціни", choices=PRICE_TYPES)
    console.print("Формат: Назва; ціна")
    while True:
        line = input("→ ").strip()
        if not line:
            break
        if ";" in line:
            name, price = [x.strip() for x in line.split(";", 1)]
            try:
                price = float(price.replace(",", "."))
            except:
                continue
            if name not in products:
                products[name] = {}
            products[name][price_type] = price
    save_json(path, products)

def mass_delete_price_type(products, path):
    pt = Prompt.ask("Тип ціни", choices=PRICE_TYPES)
    for p in products.values():
        if pt in p:
            del p[pt]
    save_json(path, products)
    console.print(f"[yellow]Ціни типу {pt} очищено[/yellow]")

def generate_invoice(company, clients, products, history, counter_path, template_path, invoice_dir):
    print_clients(clients)
    client = Prompt.ask("Оберіть клієнта").strip()
    if client not in clients:
        console.print("[red]Клієнта не знайдено[/red]")
        pause()
        return
    invoice = []
    price_type = clients[client]

    while True:
        q = Prompt.ask("Пошук товару (Enter — завершити)").strip()
        if not q:
            break
        results = context_search(products, q)
        if not results:
            console.print("[yellow]Нічого не знайдено[/yellow]")
            continue
        for i, name in enumerate(results, 1):
            console.print(f"{i}. {name} — {products[name].get(price_type, '')}")
        sel = Prompt.ask("Номер").strip()
        if not sel:
            continue
        try:
            idx = int(sel) - 1
            name = results[idx]
            qty = float(Prompt.ask("К-сть", default="1").replace(",", "."))
            price = products[name].get(price_type, 0)
            invoice.append([name, qty, price, qty * price])
        except:
            console.print("[red]Невірний вибір[/red]")

    if not invoice:
        return

    num = 1
    if os.path.exists(counter_path):
        try:
            with open(counter_path) as f:
                num = int(f.read())
        except:
            pass

    filename = f"{client}_{num:04d}.xlsx"
    filepath = os.path.join(invoice_dir, filename)
    os.makedirs(invoice_dir, exist_ok=True)

    try:
        copyfile(template_path, filepath)
        wb = load_workbook(filepath)
        ws = wb.active

        ws["B9"] = f"Рахунок №{num:04d}"
        ws["B10"] = datetime.now().strftime("%d.%m.%Y")
        ws["B7"] = client

        row = 13
        total = 0
        for item in invoice:
            ws[f"B{row}"] = item[0]
            ws[f"C{row}"] = "шт"
            ws[f"D{row}"] = item[1]
            ws[f"E{row}"] = item[2]
            ws[f"F{row}"] = item[3]
            total += item[3]
            row += 1
        ws[f"E{row}"] = total
        wb.save(filepath)

        history.setdefault(client, []).append({
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "items": invoice
        })
        with open(counter_path, "w") as f:
            f.write(str(num + 1))
        console.print(f"[green]Накладну збережено: {filepath}[/green]")
    except Exception as e:
        console.print(f"[red]Помилка: {e}[/red]")

def print_invoice_history(history):
    for client, records in history.items():
        console.print(Panel(f"Історія: {client}", style="bold cyan"))
        for r in records:
            if isinstance(r, dict) and "date" in r and "items" in r:
                console.print(f"[green]{r['date']}[/green]")
                t = Table(box=HEAVY)
                t.add_column("Назва")
                t.add_column("К-сть")
                t.add_column("Ціна")
                t.add_column("Сума")
                for it in r["items"]:
                    t.add_row(str(it[0]), str(it[1]), str(it[2]), str(it[3]))
                console.print(t)

def generate_invoice_api(company_name, client_name, items):
    data = DATA[company_name]
    customers = load_json(data["customers"])
    products = load_json(data["products"])
    history = load_json(data["history"])
    price_type = customers.get(client_name)

    if not price_type:
        raise Exception("Клієнта не знайдено або відсутній тип ціни")

    invoice = []
    for item in items:
        name = item["name"]
        qty = float(item["qty"])
        price = products.get(name, {}).get(price_type, 0)
        invoice.append([name, qty, price, qty * price])

    num = 1
    if os.path.exists(data["counter"]):
        try:
            with open(data["counter"]) as f:
                num = int(f.read())
        except:
            pass

    filename = f"{client_name}_{num:04d}.xlsx"
    filepath = os.path.join(data["invoice_dir"], filename)
    os.makedirs(data["invoice_dir"], exist_ok=True)

    copyfile(data["template"], filepath)
    wb = load_workbook(filepath)
    ws = wb.active

    ws["B9"] = f"Рахунок №{num:04d}"
    ws["B10"] = datetime.now().strftime("%d.%m.%Y")
    ws["B7"] = client_name

    row = 13
    total = 0
    for item in invoice:
        ws[f"B{row}"] = item[0]
        ws[f"C{row}"] = "шт"
        ws[f"D{row}"] = item[1]
        ws[f"E{row}"] = item[2]
        ws[f"F{row}"] = item[3]
        total += item[3]
        row += 1

    ws[f"E{row}"] = total
    wb.save(filepath)

    history.setdefault(client_name, []).append({
        "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "items": invoice
    })
    with open(data["counter"], "w") as f:
        f.write(str(num + 1))
    save_json(data["history"], history)

    return filepath

def analyze_sales(history):
    filter_str = Prompt.ask("Введіть YYYY-MM або YYYY (або Enter — всі)").strip()
    stats = {}
    for records in history.values():
        for inv in records:
            if not isinstance(inv, dict):
                continue
            if filter_str and not inv["date"].startswith(filter_str):
                continue
            for item in inv["items"]:
                stats[item[0]] = stats.get(item[0], 0) + item[1]

    if not stats:
        console.print("[yellow]Немає даних[/yellow]")
        return

    names = list(stats.keys())
    values = list(stats.values())

    plt.figure(figsize=(10, 6))
    plt.bar(names, values)
    plt.title("Продажі по товарах")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.show()

def analyze_sales_change(history):
    p1 = Prompt.ask("Введіть перший період (YYYY-MM або YYYY)").strip()
    p2 = Prompt.ask("Введіть другий період (YYYY-MM або YYYY)").strip()
    s1, s2 = {}, {}

    for records in history.values():
        for inv in records:
            if not isinstance(inv, dict):
                continue
            if inv["date"].startswith(p1):
                for item in inv["items"]:
                    s1[item[0]] = s1.get(item[0], 0) + item[1]
            elif inv["date"].startswith(p2):
                for item in inv["items"]:
                    s2[item[0]] = s2.get(item[0], 0) + item[1]

    all_items = set(s1.keys()).union(s2.keys())
    table = Table(title=f"{p1} → {p2}", box=HEAVY)
    table.add_column("Товар")
    table.add_column(p1)
    table.add_column(p2)
    table.add_column("Зміна (%)")
    for item in sorted(all_items):
        q1 = s1.get(item, 0)
        q2 = s2.get(item, 0)
        if q1 == 0 and q2 > 0:
            percent = "+∞"
        elif q1 == 0 and q2 == 0:
            percent = "-"
        else:
            percent = f"{((q2 - q1)/q1)*100:+.1f}%"
        table.add_row(item, str(q1), str(q2), percent)
    console.print(table)

def run_module(company):
    data = DATA[company]
    customers = load_json(data["customers"])
    products = load_json(data["products"])
    history = load_json(data["history"])
    os.makedirs(data["invoice_dir"], exist_ok=True)

    while True:
        console.clear()
        console.print(Panel(f"★ SUPER НАКЛАДНА ({company}) ★", width=60, style="bold cyan"))
        menu = Table(box=HEAVY)
        menu.add_row("[1] Переглянути клієнтів")
        menu.add_row("[2] Додати клієнта")
        menu.add_row("[3] Переглянути товари")
        menu.add_row("[4] Масові дії з товарами")
        menu.add_row("[5] Створити накладну")
        menu.add_row("[6] Аналіз продажів")
        menu.add_row("[7] Переглянути історію")
        menu.add_row("[0] Вихід")
        console.print(menu)

        choice = Prompt.ask("Ваш вибір", choices=[str(i) for i in range(8)])

        if choice == "1":
            print_clients(customers)
            pause()
        elif choice == "2":
            add_client(customers, data["customers"])
            pause()
        elif choice == "3":
            print_products(products)
            pause()
        elif choice == "4":
            sub = Prompt.ask("1 — Масове додавання, 2 — Очищення цін, Enter — назад").strip()
            if sub == "1":
                mass_add_products(products, data["products"])
            elif sub == "2":
                mass_delete_price_type(products, data["products"])
            pause()
        elif choice == "5":
            generate_invoice(company, customers, products, history, data["counter"], data["template"], data["invoice_dir"])
            save_json(data["history"], history)
            pause()
        elif choice == "6":
            sub = Prompt.ask("1 — Графік продажів, 2 — Порівняння періодів, Enter — назад").strip()
            if sub == "1":
                analyze_sales(history)
            elif sub == "2":
                analyze_sales_change(history)
            pause()
        elif choice == "7":
            print_invoice_history(history)
            pause()
        elif choice == "0":
            console.print("[red]Вихід[/red]")
            save_json(data["customers"], customers)
            save_json(data["products"], products)
            save_json(data["history"], history)
            break
def main():
    console.print(Panel("★ Виберіть компанію ★", style="bold cyan"))
    company = Prompt.ask("Компанія", choices=["ВФІ", "СМС"])
    run_module(company)

if __name__ == "__main__":
    main()
